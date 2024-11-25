import os
import re
import time
from flask import Flask, render_template, request, jsonify, Response, session, send_from_directory
from bs4 import BeautifulSoup
import requests
from gtts import gTTS
import openpyxl  # You're using openpyxl for Excel export
from googletrans import Translator
from googletrans.constants import LANGUAGES
from io import BytesIO

app = Flask(__name__)
app.secret_key = os.urandom(24)

translator = Translator()

tenseslist = ['Infinitif', 'Présent', 'Imparfait', 'Passé composé', 'Futur', 'Futur proche', 'Conditionnel', 'Impératif', 'Plus-que-parfait', 'Passé antérieur', 'Futur antérieur']
tenseslistbasic = ['Infinitif', 'Présent', 'Imparfait', 'Passé composé', 'Futur', 'Futur proche', 'Conditionnel', 'Impératif']

def get_word_data(word, target_lang='en'):
    reflexive_pronouns = ["me ", "te ", "se ", "s'", "nous ", "vous "]
    is_reflexive = any(word.startswith(pronoun) for pronoun in reflexive_pronouns)
    url = f"https://www.larousse.fr/dictionnaires/francais/{word}"
    headers = {'User-Agent': 'Mozilla/5.0'}
    try:
        response = requests.get(url, headers=headers, timeout=5)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, "html.parser")

        try:
            translation = translator.translate(word, src='fr', dest=target_lang)
            meaning = translation.text
        except Exception as e:
            meaning = f"Translation error: {e}"

        conjugation_url = None
        if is_reflexive:  # Special handling for reflexive verbs
            verb_link = soup.find("a", href=re.compile(r"/conjugaison/francais/se%20"))
            if verb_link:
                conjugation_url = "https://www.larousse.fr" + verb_link.get("href")
        else:
            verb_link = soup.find("a", string="Conjugaison")  # Original logic
            if verb_link:
                conjugation_url = "https://www.larousse.fr" + verb_link.get("href")
        infinitif_info = None
        definition_block = soup.find("div", id="definition")
        if definition_block:
            entry_block = definition_block.find("div", class_="Zone-Entree1")
            if entry_block:
                h2_tag = entry_block.find("h2")
                if h2_tag:
                    text_parts = [part.strip() for part in h2_tag.contents if isinstance(part, str)]
                    infinitif_info = "".join(text_parts)

        return meaning, conjugation_url, infinitif_info

    except requests.exceptions.RequestException as e:
        print(f"Request error for {word}: {e}")
        return f"Request error", None, None  # Return None for all three

    except Exception as e:
        print(f"An error occurred for {word}: {e}")
        return f"An error occurred: {e}", None, None # Return None for all three

@app.route('/set_language', methods=['POST'])
def set_language():
    target_lang = request.form.get('target_lang')
    if target_lang and (target_lang in LANGUAGES or target_lang in LANGUAGES.values()):
        session['target_lang'] = target_lang
        return jsonify({'message': 'Target language set successfully'}), 200
    else:
        return jsonify({'error': 'Invalid target language'}), 400
      
      
def get_conjugations(url):
    headers = {'User-Agent': 'Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/114.0.0.0 Safari/537.36'}
    required_tenses = tenseslist
    try:
        response = requests.get(url, headers=headers)
        response.raise_for_status()
        soup = BeautifulSoup(response.content, 'html.parser')
        conjugations = {}

        # Find Indicatif Section (if it exists)
        indicatif_section = soup.find('section', id='indicatif')
        if indicatif_section:
            for tense in required_tenses: # Iterate through ALL required tenses.
                if tense != 'Impératif': # Imperative is handled separately
                    # Find all <ul> tags within the Indicatif section
                    all_ul_tags = indicatif_section.find_all('ul')
                    for ul_tag in all_ul_tags:
                        # Check if the <ul> contains the tense name (in the <h3>)
                        h3_tag = ul_tag.find_previous_sibling('h3')
                        if h3_tag and tense in h3_tag.text:
                            conjugations[tense] = [li.text.strip() for li in ul_tag.find_all('li')]
                            break # Tense found, move to the next tense


        # Find and extract Impératif (separate section)
        imperatif_section = soup.find('section', id='imperatif')
        if imperatif_section:
            ul_tag = imperatif_section.find('ul')
            if ul_tag:
                conjugations['Impératif'] = [li.text.strip() for li in ul_tag.find_all('li')]
        # Find and extract Conditionnel (separate section)
        conditionnel_section = soup.find('section', id='conditionnel')
        if conditionnel_section:
            ul_tag = conditionnel_section.find('ul')
            if ul_tag:
                conjugations['Conditionnel'] = [li.text.strip() for li in ul_tag.find_all('li')]
                
        return conjugations

    except requests.exceptions.RequestException as e:
        print(f"Request error: {e}")
        return {}
    except AttributeError as e:
        print(f"AttributeError during scraping: {e}, URL: {url}")
        return {}
    except Exception as e:
        print(f"An error occurred: {e}")
        return {}

@app.route('/', methods=['GET', 'POST'])
def index():
    if request.method == 'GET':
        session.pop('results', None)  # Clear results on GET request
        return render_template('index.html', target_lang=session.get('target_lang', 'en'))

    elif request.method == 'POST':
        session.pop('results', None)
        target_lang = session.get('target_lang', 'en')
        verbs = request.form.getlist('verbs[]')
        if not verbs or all(v.strip() == '' for v in verbs):
            return jsonify({'error': 'No verbs provided for translation'}), 400

        results = []
        for verb in verbs:
            reflexive_pronouns = ["me ", "te ", "se ", "s'", "nous ", "vous "]
            is_reflexive = any(verb.startswith(pronoun) for pronoun in reflexive_pronouns)
            verb = verb.strip()
            if verb:
                meaning, conjugation_url, infinitif_info = get_word_data(verb, target_lang)

                conjugations = {}  # Important: Initialize conjugations HERE

                if conjugation_url:
                    conjugations = get_conjugations(conjugation_url)

                #Correctly handle the infinitive even if it's pronominal
                if is_reflexive and 'Présent' in conjugations:  #Handle reflexive infinitives not in main infinitive block
                    conjugations['Infinitif'] = ["se " + infinitif_info]
                elif infinitif_info:
                    conjugations['Infinitif'] = [infinitif_info]
                elif 'Infinitif' in conjugations and conjugations['Infinitif']:
                    conjugations['Infinitif'] = [conjugations['Infinitif'][0]] #Clean up the infinitive
                    
                #Improved Futur proche handling:
                if 'Présent' in conjugations and conjugations['Infinitif']:
                    if is_reflexive:
                        pronouns = ["je vais me", "tu vas te", "il, elle va se", "nous allons nous", "vous allez vous", "ils, elles vont se"]
                        conjugations['Futur proche'] = [f"{pronoun} {infinitif_info}" for pronoun in pronouns]
                    else:  # Non-reflexive verbs
                        conjugations['Futur proche'] = [
                            f"{pronoun} {infinitif_info}"
                            for pronoun in ["je vais", "tu vas", "il, elle va", "nous allons", "vous allez", "ils, elles vont"]
                        ]
                results.append({'verb': verb, 'meaning': meaning, 'conjugations': conjugations})

        session['results'] = results
        return jsonify({'results': results})  # Return JSON after processing

    return render_template('index.html', target_lang=session.get('target_lang', 'en'))  # Handle other cases if necessary
  
@app.route('/pronounce', methods=['POST'])
def pronounce():
    try:
        text = request.form.get('text')
        lang = request.form.get('lang', 'fr')

        processed_texts = preprocess_french_verb(text)

        if isinstance(processed_texts, list):  # Handle multiple pronunciations (e.g., masculine/feminine)
            audio_data = BytesIO()  # Use a single BytesIO object to combine audio
            for text_to_pronounce in processed_texts:
                tts = gTTS(text=text_to_pronounce, lang=lang)
                tts.write_to_fp(audio_data) # Append audio to existing BytesIO
            audio_data.seek(0)
            response = Response(audio_data.read(), mimetype='audio/mpeg')

        else:  # Handle single pronunciation
            tts = gTTS(text=processed_texts, lang=lang)
            fp = BytesIO()
            tts.write_to_fp(fp)
            fp.seek(0)
            response = Response(fp.read(), mimetype='audio/mpeg')


        return response


    except Exception as e:
        import traceback
        print(f"Error in /pronounce: {e}")
        traceback.print_exc()
        return jsonify({'error': str(e)}), 500

def preprocess_french_verb(text):
    """
    Preprocesses French verb conjugations based on the provided logic.
    """
    pronoun_match = re.match(r"^(je|tu|il, elle|nous|vous|ils, elles|on)\s+(.+)", text)

    if not pronoun_match:
        return text  # No pronoun found, return original text

    pronoun = pronoun_match.group(1)
    verb = pronoun_match.group(2)

    results = []

    if "(e)s" in verb:
        verb_stem = verb.replace("(e)s", "") # remove the complete (e)s
        if pronoun == "il, elle":
            results.extend([f"il {verb_stem}s", f"elle {verb_stem}es"])  # Correct singular/plural
        elif pronoun == "ils, elles":
            results.extend([f"ils {verb_stem}s", f"elles {verb_stem}es"]) #correct singular/plural
        else:
            results.extend([f"{pronoun} {verb_stem}s", f"{pronoun} {verb_stem}es"])
            
    
    elif "(e)" in verb:
        verb_stem = verb.replace("(e)", "")
        if pronoun == "il, elle":
            results.extend([f"il {verb_stem}", f"elle {verb_stem}e"])
        elif pronoun == "ils, elles":
            results.extend([f"ils {verb_stem}s", f"elles {verb_stem}es"])
        else:
            results.extend([f"{pronoun} {verb_stem}", f"{pronoun} {verb_stem}e"])


    elif "/" in verb:
        parts = verb.split("/")
        if pronoun == "il, elle":
            results.extend([f"il {parts[0].strip()}", f"elle {parts[1].strip()}"])
        elif pronoun == "ils, elles":
            results.extend([f"ils {parts[0].strip()}", f"elles {parts[1].strip()}"])
        else:
            results.extend([f"{pronoun} {parts[0].strip()}", f"{pronoun} {parts[1].strip()}"])

    else:
        results.append(f"{pronoun} {verb}")  # No transformations needed
    return results

@app.route('/export', methods=['POST'])
def export_excel():
    results = session.get('results') #Retrieve results from the session

    if not results:
        return jsonify({'error': 'No data to export. Please submit verbs first.'}), 400
    
    advance_mode = request.headers.get('X-Advance-Mode')  # Get the mode from the header
    tenseslistexport = tenseslist if advance_mode == 'Advance' else tenseslistbasic
    workbook = openpyxl.Workbook()
    sheet = workbook.active
    sheet.append(['Infinitif', 'Traduction'] + tenseslistexport)

    for row_index, result in enumerate(results, 1):  # Start from 1 for row indexing
        sheet.cell(row=row_index + 1, column=1, value=result['verb'])  # Verb in column A
        sheet.cell(row=row_index + 1, column=2, value=result['meaning'])  # Meaning in column B

        for col_index, tense in enumerate(tenseslistexport, 3):  # Start from 3 for column indexing
            conj_list = result['conjugations'].get(tense, [])
            print(result['conjugations'])
            cell_value = "\n".join(conj_list) if conj_list else ""
            cell = sheet.cell(row=row_index + 1, column=col_index, value=cell_value)
            cell.alignment = openpyxl.styles.Alignment(wrapText=True)
            if conj_list:
                sheet.row_dimensions[row_index + 1].height = len(conj_list) * 15
                
    output = BytesIO() #Create a BytesIO object to hold the Excel file data
    workbook.save(output)
    output.seek(0) 

    try:
        timestr = time.strftime("%Y%m%d-%H%M%S")
        filename = f"verb_conjugations_{timestr}.xlsx"
        response = Response(
        output,
        mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",  # Correct MIME type
        headers={"Content-Disposition": f"attachment; filename={filename}"}
    )
        return response
    except Exception as e:
        return jsonify({'error': str(e)}), 500

if __name__ == '__main__':
    app.run()
